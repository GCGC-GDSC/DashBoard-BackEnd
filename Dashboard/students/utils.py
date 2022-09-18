from rest_framework import generics, status, views, response
from organization.models import Institute, Campus, Stream, Programs
from organization.serializers import CampusSerialize, InstituteSerialize
from rest_framework.response import Response
from rest_framework.parsers import FileUploadParser
from django.http import JsonResponse
from rest_framework import generics
from django.http import HttpResponse
from wsgiref.util import FileWrapper
from tablib import Dataset
from rest_framework.decorators import api_view
from rest_framework.status import *
from .serializers import *
from .models import *
import pandas as pd
import openpyxl
import datetime
import calendar
import traceback
import logging
import os
"""
class FileUploadView(views.APIView):
    parser_classes = [FileUploadParser]

    def put(self, request, format=None):
        excel_file = request.data['file']
        print(excel_file)
        dataset = Dataset()
        if not excel_file.name.endswith('.xlsx'):
            return Response("File should be excel only", status=404)
        imported_data = dataset.load(excel_file.read(),
                                     headers=False,
                                     format='xlsx')
        print("imported_data: \n", imported_data)

        data = {}
        try:
            under_campus = Campus.objects.get(name=imported_data[0][1])
            under_institute = Institute.objects.get(name=imported_data[1][1])
            is_ug = imported_data[-1][1]
        except Campus.DoesNotExist:
            return Response("Campus Invalid: " + str(imported_data))
        except Institute.DoesNotExist:
            return Response("Institute Invalid" + str(imported_data[1]))

        for key_val in imported_data[2:-1]:
            data[key_val[0]] = key_val[1]
        qs, create = Graduates.objects.get_or_create(
            under_campus=under_campus,
            under_institute=under_institute,
            is_ug=is_ug)
        qs.total_students = data['total_students']
        qs.total_final_years = data['total_final_years']
        qs.total_higher_study_and_pay_crt = data[
            'total_higher_study_and_pay_crt']
        qs.total_not_intrested_in_placments = data[
            'total_not_intrested_in_placments']
        qs.total_backlogs_opted_for_placements = data[
            'total_backlogs_opted_for_placements']
        qs.total_backlogs_opted_for_higherstudies = data[
            'total_backlogs_opted_for_higherstudies']
        qs.total_backlogs_opted_for_other_career_options = data[
            'total_backlogs_opted_for_other_career_options']
        qs.total_offers = data['total_offers']
        qs.total_multiple_offers = data['total_multiple_offers']
        qs.highest_salary = data['highest_salary']
        qs.lowest_salary = data['lowest_salary']
        qs.average_salary = data['average_salary']
        qs.save()

        return Response("Data sent", status=204)

    def post(self, request, format=None):
        excel_file = request.data['file']
        dataset = Dataset()
        if not excel_file.name.endswith('.xlsx'):
            return Response("File should be excel only", status=404)
        imported_data = dataset.load(excel_file.read(),
                                     headers=False,
                                     format='xlsx')

        data = {}

        try:
            under_campus = Campus.objects.get(name=imported_data[0][1])
            under_institute = Institute.objects.get(name=imported_data[1][1])
        except Campus.DoesNotExist:
            return Response("Campus Invalid: " + str(imported_data))
        except Institute.DoesNotExist:
            return Response("Institute Invalid" + str(imported_data[1]))

        for key_val in imported_data[2:]:
            data[key_val[0]] = key_val[1]

        qs = Graduates(under_campus=under_campus,
                       under_institute=under_institute,
                       **data)
        qs.save()
        return Response("Data sent", status=204)
"""


def LastNlines(fname, N=10):

    with open(fname) as f:
        log_buffer = f.readlines()
        log_buffer = log_buffer[-1:-N:-1]

    return log_buffer


@api_view(('GET', ))
def log_edit_info(request):
    db_logger = logging.getLogger('db')
    try:
        # with open("./logs/dblog.txt", "r") as file:
        #     i = 0
        #     lines_size = 10
        #     last_lines = []
        #     for line in file:
        #         if i < lines_size:
        #             last_lines.append(line)
        #         else:
        #             last_lines[i % lines_size] = line
        #         i = i + 1

        # last_lines = last_lines[(i % lines_size):] + last_lines[:(i %
        #                                                           lines_size)]

        # send_data = []

        # for line in last_lines:
        #     send_data.append(line)
        send_data = LastNlines(fname="./logs/dblog.txt")
        return Response({'status': 'ok', 'result': send_data})
    except Exception as e:
        db_logger.exception(traceback.print_exc())
        return Response({'status': 'Error', 'result': str(e)})


def export_data_to_excel(request, name, year):
    ext = '.xlsx'
    searchfilename = name + ext
    if name.lower() == 'overall':
        obj = Graduates.objects.filter(passing_year=year)
    elif name.lower() == 'gst':
        camp = Campus.objects.get(name='vskp')
        inst = Institute.objects.get(name='gst', under_campus=camp)
        obj = GraduatesWithPrograms.objects.filter(passing_year=year,
                                                   under_institute=inst,
                                                   under_campus=camp)
        print("objects: ", obj)
    else:
        camp = Campus.objects.get(name=name)
        obj = Graduates.objects.filter(passing_year=year, under_campus=camp)
    # print("All objects: ", obj)
    data = []

    for i in obj:
        if i.total_final_years == 0:
            Percentage_of_students_opted_HS_to_the_total_number = 0
            Percentage_of_students_having_backlogs_to_the_total_number_of_students = 0
            Percentage_of_students_eligible_for_and_requiring_placement = 0
        else:
            Percentage_of_students_opted_HS_to_the_total_number = round((
                (i.total_opted_for_higher_studies_only / i.total_final_years) *
                100), 2)
            Percentage_of_students_having_backlogs_to_the_total_number_of_students = round(
                ((i.total_backlogs / i.total_final_years) * 100), 2)
            Percentage_of_students_eligible_for_and_requiring_placement = round(
                ((i.total_students_eligible / i.total_final_years) * 100), 2)

        if i.total_students_eligible == 0:
            Percentage_of_students_placed_out_of_eligible_students = 0
        else:
            Percentage_of_students_placed_out_of_eligible_students = round(
                ((i.total_placed / i.total_students_eligible) * 100), 2)

        if i.total_students_eligible == 0:
            Percentage_of_students_yet_to_be_placed_out_of_eligible_students = 0
        else:
            Percentage_of_students_yet_to_be_placed_out_of_eligible_students = round(
                ((i.total_yet_to_place / i.total_students_eligible) * 100), 2)
        program = ""
        try:
            program = i.program
        except:
            pass

        data.append({
            "total_students":
            i.total_students,
            "total_final_years":
            i.total_final_years,
            "total_higher_study_and_pay_crt":
            i.total_higher_study_and_pay_crt,
            "total_opted_for_higher_studies_only":
            i.total_opted_for_higher_studies_only,
            "No_of_students_opted_out_of_any_Career_Fulfillment_activities":
            i.total_not_intrested_in_placments,
            "total_backlogs":
            i.total_backlogs,
            "total_backlogs_opted_for_placements":
            i.total_backlogs_opted_for_placements,
            "total_backlogs_opted_for_higherstudies":
            i.total_backlogs_opted_for_higherstudies,
            "total_backlogs_opted_for_other_career_options":
            i.total_backlogs_opted_for_other_career_options,
            "total_students_eligible":
            i.total_students_eligible,
            "total_offers":
            i.total_offers,
            "total_multiple_offers":
            i.total_multiple_offers,
            "total_placed":
            i.total_placed,
            "total_yet_to_place":
            i.total_yet_to_place,
            "Percentage_of_students_opted_HS_to_the_total_number":
            Percentage_of_students_opted_HS_to_the_total_number,
            "Percentage_of_students_having_backlogs_to_the_total_number_of_students":
            Percentage_of_students_having_backlogs_to_the_total_number_of_students,
            "Percentage_of_students_eligible_for_and_requiring_placement":
            Percentage_of_students_eligible_for_and_requiring_placement,
            "Percentage_of_students_placed_out_of_eligible_students":
            Percentage_of_students_placed_out_of_eligible_students,
            "Percentage_of_students_yet_to_be_placed_out_of_eligible_students":
            Percentage_of_students_yet_to_be_placed_out_of_eligible_students,
            "salary_details":
            "",
            "highest_salary":
            i.highest_salary,
            "lowest_salary":
            i.lowest_salary,
            "average_salary":
            i.average_salary,
            "is_ug":
            i.is_ug,
            "under_institute_name":
            i.under_institute_name,
            "under_campus_name":
            i.under_campus_name,
            "under_campus":
            i.under_campus,
            "under_institute":
            i.under_institute,
            "program":
            program
        })

    # print("============================================")
    # print(data)
    # print("data: ", len(data))
    # print("============================================")
    searchpath = "media/" + searchfilename
    #print("=============================", searchpath, "============================")
    wb = openpyxl.load_workbook(searchpath)
    if name.lower() == 'gst':
        sheet = wb.get_sheet_by_name('CF 2022')
        print("sheet ==>>", sheet)
        sheet_obj = wb.active
        dic = {}

        for x in range(3, sheet_obj.max_column + 1):
            val = (sheet_obj.cell(row=3, column=x).value)
            if val == None:
                continue
            dic[val.lower()] = x

        print("dic value: =====>", dic)

        for da in data:
            inst = da['program']
            try:
                num = 5
                val = dic[inst.lower()]
                #print("val", val, inst)
                if da['is_ug'] is False:
                    val += 1

                for i in da:
                    if num < 28:
                        totnum = 64 + val
                        if totnum <= 90:
                            cell = chr(64 + val) + str(num)
                        else:
                            diff = totnum - 90
                            cell = chr(65) + chr(64 + diff) + str(num)
                        inpval = da[i]
                        sheet[cell] = inpval
                    num += 1
            except:
                print("does not belong to this campus", inst)
        wb.save('media/out.xlsx')
        return JsonResponse({'status': 200})
    else:
        sheet = wb.get_sheet_by_name('CF 2022')

    sheet_obj = wb.active
    dic = {}

    # if name.lower()=='overall':
    #     try:
    #         for da in data:
    #             for i in da:
    #                 pass
    #     except:
    #         print("does not belong to this campus: ", inst)

    # camp_vals = []

    campus_id_values = {
        'Visakhapatnam Campus': 'vskp',
        'Hyderabad Campus': 'hyd',
        'Bengaluru Campus': 'blr',
    }

    # for x in range(3, sheet_obj.max_column + 1):
    #     val = (sheet_obj.cell(row=2, column=x).value)
    #     if val == None:
    #         continue
    #     if val.lower() == 'total':
    #         continue
    #     camp_vals.append(campus_id_values[val.lower()])

    # print("campvals: ", camp_vals)

    for x in range(3, sheet_obj.max_column + 1):
        val = (sheet_obj.cell(row=3, column=x).value)
        camp_val = (sheet_obj.cell(row=2, column=x).value)
        if camp_val == None and val == None:
            continue
        if camp_val != None:
            if camp_val != 'Total':
                camp_val = campus_id_values[camp_val]
                dic[camp_val] = {}
            if val != None:
                tempval = list(dic)[-1]
                dic[tempval][val.lower()] = x

        elif dic != {} and val != None:
            tempval = list(dic)[-1]
            # print("temp_Val: ", tempval)
            dic[tempval][val.lower()] = x

    print("this is the value: ", dic)

    for da in data:
        inst = da['under_institute_name']
        camp = da['under_campus']
        # print(camp, "-->", inst)
        # print("values: ", dic[camp.name][inst])
        try:
            num = 5
            val = dic[camp.name][inst.lower()]
            # print("===>>", val)
            #print("val", val, inst)
            if da['is_ug'] is False:
                val += 1
            for i in da:
                if num < 28:
                    totnum = 64 + val
                    if totnum <= 90:
                        cell = chr(64 + val) + str(num)
                    else:
                        diff = totnum - 90
                        cell = chr(65) + chr(64 + diff) + str(num)
                    inpval = da[i]
                    sheet[cell] = inpval
                num += 1
        except:
            print("does not belong to this campus", inst)
    wb.save('media/out.xlsx')
    return JsonResponse({'status': 200})


class FileDownloadListAPIView(generics.ListAPIView):
    serializer_class = GraduatesSerializer

    def get(self, request, year, name, format=None):
        db_logger = logging.getLogger('db')
        try:
            filename = f"{str(name).upper()} Career Fulfillment Statistics - 2022 Batch"
            try:
                export_data_to_excel(request, name, year)
            except Exception as e:
                return Response({
                    'status': 'error',
                    'result': str(e)
                },
                                status=status.HTTP_501_NOT_IMPLEMENTED)

            document = open('media/out.xlsx', 'rb')
            filename = filename + '.xlsx'
            response = HttpResponse(FileWrapper(document),
                                    content_type='application/msexcel')
            response[
                'Content-Disposition'] = 'attachment; filename="%s"' % filename
            return response
        except Exception as e:
            db_logger.exception(traceback.print_exc())
